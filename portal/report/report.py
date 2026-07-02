#!/usr/bin/env python3
"""
Omnignis church livestream report generator.

Runs on a schedule (GitHub Actions, daily). For every church whose report is
"due" today (daily / weekly = Sunday / monthly = 1st), it:
  1. decrypts that church's Facebook page token,
  2. pulls the videos since the last report,
  3. reads total views + unique viewers for each,
  4. builds a formatted Excel workbook,
  5. emails it to the church's destination address(es) via Resend,
  6. records the run time.

NOTE ON FACEBOOK METRICS: Facebook's Graph API changes over time. The metric
names and edges below (video_insights: total_video_views, total_video_views_unique)
are correct as of the current API but should be verified against a real page
during your first test run, and adjusted here if Facebook returns an error.
"""

import os
import io
import base64
import hmac
import hashlib
from datetime import datetime, timedelta, timezone

import requests
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUPABASE_URL = os.environ["SUPABASE_URL"].rstrip("/")
SERVICE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
ENC_KEY = base64.b64decode(os.environ["TOKEN_ENCRYPTION_KEY"])
APP_SECRET = os.environ.get("FACEBOOK_APP_SECRET", "")
GRAPH_VERSION = os.environ.get("GRAPH_API_VERSION", "v21.0")
RESEND_API_KEY = os.environ["RESEND_API_KEY"]
FROM_EMAIL = os.environ.get("REPORT_FROM_EMAIL", "reports@omnignis.com")

GRAPH = f"https://graph.facebook.com/{GRAPH_VERSION}"
SB_HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
}


# ----------------------- helpers -----------------------
def decrypt_token(b64: str) -> str:
    raw = base64.b64decode(b64)
    iv, ct_tag = raw[:12], raw[12:]          # matches lib/crypto.js layout
    return AESGCM(ENC_KEY).decrypt(iv, ct_tag, None).decode("utf-8")


def appsecret_proof(token: str) -> str:
    if not APP_SECRET:
        return ""
    return hmac.new(APP_SECRET.encode(), token.encode(), hashlib.sha256).hexdigest()


def sb_get(table: str, params: str = "") -> list:
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}?{params}", headers=SB_HEADERS, timeout=30)
    r.raise_for_status()
    return r.json()


def sb_patch(table: str, match: str, body: dict) -> None:
    r = requests.patch(f"{SUPABASE_URL}/rest/v1/{table}?{match}", headers=SB_HEADERS, json=body, timeout=30)
    r.raise_for_status()


def is_due(frequency: str, today: datetime) -> bool:
    if frequency == "daily":
        return True
    if frequency == "weekly":
        return today.weekday() == 6          # Sunday
    if frequency == "monthly":
        return today.day == 1
    return False


def lookback_days(frequency: str) -> int:
    return {"daily": 2, "weekly": 8, "monthly": 32}.get(frequency, 8)


# ----------------------- facebook -----------------------
def fetch_videos(page_id: str, token: str, since_dt: datetime) -> list:
    proof = appsecret_proof(token)
    params = {
        "access_token": token,
        "fields": "id,title,description,created_time",
        "since": int(since_dt.timestamp()),
        "limit": 50,
    }
    if proof:
        params["appsecret_proof"] = proof
    videos, url = [], f"{GRAPH}/{page_id}/videos"
    while url and len(videos) < 200:
        r = requests.get(url, params=params, timeout=30)
        data = r.json()
        if "error" in data:
            raise RuntimeError(f"Graph videos error: {data['error'].get('message')}")
        videos.extend(data.get("data", []))
        url = data.get("paging", {}).get("next")
        params = None                     # "next" URLs already carry the query string
    return videos


def fetch_metrics(video_id: str, token: str) -> dict:
    proof = appsecret_proof(token)
    params = {
        "access_token": token,
        "metric": "total_video_views,total_video_views_unique",
    }
    if proof:
        params["appsecret_proof"] = proof
    r = requests.get(f"{GRAPH}/{video_id}/video_insights", params=params, timeout=30)
    data = r.json()
    out = {"total_views": 0, "unique_viewers": 0}
    for item in data.get("data", []):
        name = item.get("name")
        try:
            value = item["values"][0]["value"]
        except (KeyError, IndexError):
            value = 0
        if name == "total_video_views":
            out["total_views"] = value
        elif name == "total_video_views_unique":
            out["unique_viewers"] = value
    return out


# ----------------------- excel -----------------------
def build_workbook(church_name: str, rows: list, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Livestream Report"

    ember = "FF6A1A"
    dark = "0D0F14"
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:D1")
    ws["A1"] = church_name
    ws["A1"].font = Font(size=16, bold=True, color=dark)
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Livestream attendance report  ·  {period_label}"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    headers = ["Livestream Date", "Title", "Total Views", "Unique Viewers"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, 5):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ember)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    total_v = total_u = 0
    for row in rows:
        ws.append([row["date"], row["title"], row["total_views"], row["unique_viewers"]])
        total_v += row["total_views"] or 0
        total_u += row["unique_viewers"] or 0
        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).border = border

    ws.append(["", "Total", total_v, total_u])
    for col in range(1, 5):
        c = ws.cell(row=ws.max_row, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F0F0F0")
        c.border = border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------- email -----------------------
def send_email(to_list: list, subject: str, html: str, filename: str, content: bytes) -> None:
    payload = {
        "from": f"Omnignis Reports <{FROM_EMAIL}>",
        "to": to_list,
        "subject": subject,
        "html": html,
        "attachments": [{"filename": filename, "content": base64.b64encode(content).decode()}],
    }
    r = requests.post(
        "https://api.resend.com/emails",
        headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
        json=payload, timeout=30,
    )
    if r.status_code >= 300:
        raise RuntimeError(f"Resend error {r.status_code}: {r.text}")


# ----------------------- main -----------------------
def main():
    today = datetime.now(timezone.utc)
    profiles = sb_get("profiles", "select=id,church_name,destination_emails,report_frequency,last_report_at")
    connections = {
        c["profile_id"]: c
        for c in sb_get("facebook_connections", "select=profile_id,page_id,page_name,token_ciphertext")
        if c.get("page_id") and c.get("token_ciphertext")   # skip half-finished connections
    }

    processed = 0
    for p in profiles:
        conn = connections.get(p["id"])
        if not conn:
            continue
        if not is_due(p.get("report_frequency", "weekly"), today):
            continue

        try:
            token = decrypt_token(conn["token_ciphertext"])
            if p.get("last_report_at"):
                since = datetime.fromisoformat(p["last_report_at"].replace("Z", "+00:00"))
            else:
                since = today - timedelta(days=lookback_days(p["report_frequency"]))

            videos = fetch_videos(conn["page_id"], token, since)
            rows = []
            for vid in videos:
                m = fetch_metrics(vid["id"], token)
                created = vid.get("created_time", "")[:10]
                rows.append({
                    "date": created,
                    "title": vid.get("title") or vid.get("description", "")[:60] or "Livestream",
                    "total_views": m["total_views"],
                    "unique_viewers": m["unique_viewers"],
                })

            period = f"{since.date()} to {today.date()}"
            xlsx = build_workbook(p["church_name"], rows, period)
            recipients = [e.strip() for e in (p["destination_emails"] or "").split(",") if e.strip()]
            if not recipients:
                continue

            html = (
                f"<p>Hello {p['church_name']},</p>"
                f"<p>Attached is your livestream attendance report for <b>{period}</b> "
                f"({len(rows)} livestream{'s' if len(rows) != 1 else ''}).</p>"
                f"<p>— Omnignis Technologies</p>"
            )
            fname = f"{p['church_name'].replace(' ', '_')}_livestream_report_{today.date()}.xlsx"
            send_email(recipients, f"Livestream report — {p['church_name']}", html, fname, xlsx)

            sb_patch("profiles", f"id=eq.{p['id']}", {"last_report_at": today.isoformat()})
            processed += 1
            print(f"OK: sent report to {p['church_name']} ({len(rows)} videos)")
        except Exception as e:
            print(f"ERROR for {p.get('church_name')}: {e}")

    print(f"Done. Reports sent: {processed}")


if __name__ == "__main__":
    main()
